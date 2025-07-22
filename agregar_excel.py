from flask import Flask, render_template_string, request, redirect
from openpyxl import load_workbook

app = Flask(__name__)

archivo_excel = 'REPORTE.xlsx'

html_form = '''
<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="UTF-8">
  <title>Revalidaciones Manual</title>
  <style>
    body { font-family: Arial, sans-serif; background: #f0f2f5; padding: 20px; }
    h2, h3 { text-align: center; }
    form { max-width: 600px; margin: auto; display: flex; gap: 10px; flex-wrap: wrap; margin-bottom: 20px; }
    input, select, button { padding: 8px; font-size: 14px; border-radius: 5px; border: 1px solid #ccc; flex: 1; }
    button, input[type=submit] { background-color: #007BFF; color: white; cursor: pointer; }
    .contenedor-lista { display: flex; flex-direction: column; gap: 15px; max-width: 600px; margin: auto; }
    .cuadro { background-color: white; border-left: 5px solid #007BFF; border-radius: 8px; padding: 15px; box-shadow: 0 2px 6px rgba(0,0,0,0.1); }
    .cuadro h4 { margin: 0 0 10px; }
    .estado { font-weight: bold; }
    .revalidado { color: green; }
    .programado { color: goldenrod; }
    .pendiente { color: red; }
  </style>
</head>
<body>

  <h2>Agregar Revalidación</h2>

  <form method="POST">
    <input type="text" name="referencia" placeholder="REFERENCIA" required>
    <input type="text" name="mbl" placeholder="MBL" required>
    <select name="estatus" required>
      <option value="" disabled selected>-- Seleccione Estado --</option>
      <option value="REVALIDADO">REVALIDADO</option>
      <option value="PROGRAMADO">PROGRAMADO</option>
      <option value="PENDIENTE">PENDIENTE</option>
    </select>
    <input type="submit" value="Agregar">
  </form>

  <form method="GET">
    <select name="filtro">
      <option value="">-- Filtrar por Estado --</option>
      <option value="REVALIDADO" {% if filtro == 'REVALIDADO' %}selected{% endif %}>REVALIDADO</option>
      <option value="PROGRAMADO" {% if filtro == 'PROGRAMADO' %}selected{% endif %}>PROGRAMADO</option>
      <option value="PENDIENTE" {% if filtro == 'PENDIENTE' %}selected{% endif %}>PENDIENTE</option>
    </select>
    <button type="submit">Filtrar</button>
  </form>

  <h3>Registros en Excel:</h3>
  <div class="contenedor-lista">
    {% for fila in datos %}
      {% set clase = "" %}
      {% if fila[2] == "REVALIDADO" %}
        {% set clase = "revalidado" %}
      {% elif fila[2] == "PROGRAMADO" %}
        {% set clase = "programado" %}
      {% elif fila[2] == "PENDIENTE" %}
        {% set clase = "pendiente" %}
      {% endif %}
      <div class="cuadro">
        <h4><strong>Referencia:</strong> {{ fila[0] }}</h4>
        <p><strong>BL:</strong> {{ fila[1] }}</p>
        <p><strong>Estado:</strong> <span class="estado {{ clase }}">{{ fila[2] }}</span></p>
      </div>
    {% endfor %}
  </div>

</body>
</html>
'''

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        referencia = request.form.get('referencia')
        mbl = request.form.get('mbl')
        estatus = request.form.get('estatus')

        wb = load_workbook(archivo_excel)
        ws = wb.active

        fila = ws.max_row + 1
        ws.cell(row=fila, column=1, value=referencia)
        ws.cell(row=fila, column=2, value=mbl)
        ws.cell(row=fila, column=3, value=estatus)

        wb.save(archivo_excel)
        return redirect('/')

    # Aplicar filtro si existe
    filtro = request.args.get('filtro', '')
    wb = load_workbook(archivo_excel)
    ws = wb.active
    datos = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not filtro or (row[2] == filtro):
            datos.append(row)

    return render_template_string(html_form, datos=datos, filtro=filtro)

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=8080)

